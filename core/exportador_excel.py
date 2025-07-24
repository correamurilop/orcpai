from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from core.regras_engine import aplicar_regras_orcamento

def exportar_orcamento_para_excel(orcamento, caminho_arquivo: str):
    """
    Gera uma planilha .xlsx contendo todos os componentes de um orçamento,
    incluindo os derivados pelas regras, organizados por painel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento de Painéis"

    # Estilos
    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    derivado_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    center_alignment = Alignment(horizontal="center")

    # Título da planilha
    ws.merge_cells('A1:D1')
    ws['A1'] = f"ORÇAMENTO: {orcamento.nome}"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center_alignment

    # Informações do orçamento
    ws['A2'] = f"Data: {orcamento.data}"
    ws['A3'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

    # Cabeçalhos
    row = 5
    headers = ["Painel", "Componente", "Quantidade", "Tipo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    row += 1

    # Componentes diretos
    for painel in orcamento.paineis:
        for componente in painel.componentes:
            ws.cell(row=row, column=1, value=painel.nome)
            ws.cell(row=row, column=2, value=componente.nome)
            ws.cell(row=row, column=3, value=componente.quantidade)
            ws.cell(row=row, column=4, value="Direto")
            row += 1

    # Separador
    ws.cell(row=row, column=1, value="")
    row += 1
    
    # Título para componentes derivados
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="COMPONENTES GERADOS AUTOMATICAMENTE (REGRAS)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = center_alignment
    row += 1

    # Componentes derivados por regras
    componentes_derivados = aplicar_regras_orcamento(orcamento)
    
    if componentes_derivados:
        for painel_nome, componente_derivado in componentes_derivados:
            cell_painel = ws.cell(row=row, column=1, value=painel_nome)
            cell_comp = ws.cell(row=row, column=2, value=componente_derivado.nome)
            cell_qtd = ws.cell(row=row, column=3, value=componente_derivado.quantidade)
            cell_tipo = ws.cell(row=row, column=4, value="Gerado")
            
            # Destacar componentes derivados
            for cell in [cell_painel, cell_comp, cell_qtd, cell_tipo]:
                cell.fill = derivado_fill
            
            row += 1
    else:
        ws.cell(row=row, column=1, value="Nenhum componente gerado automaticamente")
        ws.cell(row=row, column=1).font = Font(italic=True)
        row += 1

    # Resumo
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="RESUMO")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = center_alignment
    row += 1

    # Contar painéis e componentes
    total_paineis = len(orcamento.paineis)
    total_componentes_diretos = sum(len(p.componentes) for p in orcamento.paineis)
    total_componentes_derivados = len(componentes_derivados)

    ws.cell(row=row, column=1, value="Total de Painéis:")
    ws.cell(row=row, column=2, value=total_paineis)
    row += 1

    ws.cell(row=row, column=1, value="Componentes Diretos:")
    ws.cell(row=row, column=2, value=total_componentes_diretos)
    row += 1

    ws.cell(row=row, column=1, value="Componentes Gerados:")
    ws.cell(row=row, column=2, value=total_componentes_derivados)
    row += 1

    ws.cell(row=row, column=1, value="TOTAL GERAL:")
    ws.cell(row=row, column=2, value=total_componentes_diretos + total_componentes_derivados)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)

    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salvar arquivo
    wb.save(caminho_arquivo)

def exportar_resumo_componentes(orcamento, caminho_arquivo: str):
    """
    Gera uma planilha Excel com resumo consolidado de todos os componentes
    (agrupados por nome, somando quantidades)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo de Componentes"

    # Consolidar todos os componentes
    componentes_consolidados = {}
    
    # Componentes diretos
    for painel in orcamento.paineis:
        for componente in painel.componentes:
            if componente.nome in componentes_consolidados:
                componentes_consolidados[componente.nome] += componente.quantidade
            else:
                componentes_consolidados[componente.nome] = componente.quantidade
    
    # Componentes derivados
    derivados = aplicar_regras_orcamento(orcamento)
    for painel_nome, componente_derivado in derivados:
        nome_completo = f"{componente_derivado.nome} (gerado)"
        if nome_completo in componentes_consolidados:
            componentes_consolidados[nome_completo] += componente_derivado.quantidade
        else:
            componentes_consolidados[nome_completo] = componente_derivado.quantidade

    # Estilos
    titulo_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Título
    ws.merge_cells('A1:B1')
    ws['A1'] = f"RESUMO CONSOLIDADO - {orcamento.nome}"
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal="center")

    # Cabeçalhos
    ws['A3'] = "Componente"
    ws['B3'] = "Quantidade Total"
    
    for cell in [ws['A3'], ws['B3']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Dados
    row = 4
    for nome, quantidade in sorted(componentes_consolidados.items()):
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=quantidade)
        row += 1

    # Ajustar colunas
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15

    wb.save(caminho_arquivo) 